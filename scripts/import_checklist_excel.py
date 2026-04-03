#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "Pentest_Checklist_v2.xlsx"
DEFAULT_OUT = ROOT / "data" / "checklistCatalog.import.json"

SHEET_MAP = {
    "WEB - Baseline": ("web", "baseline"),
    "WEB - Custom": ("web", "custom"),
    "MOBILE - Baseline": ("mobile", "baseline"),
    "MOBILE - Custom": ("mobile", "custom"),
    "DESKTOP - Baseline": ("desktop", "baseline"),
    "DESKTOP - Custom": ("desktop", "custom"),
}


def slug(platform: str, label: str | None) -> str | None:
    if not label:
        return None
    token = re.sub(r"[^a-z0-9]+", "-", label.lower()).strip("-")
    return f"{platform}:{token}"


def normalize_access(mode: str) -> str:
    return {
        "Black-Box": "blackbox",
        "Scoped": "greybox",
        "Grey-Box": "greybox",
        "Both": "both",
    }.get(mode or "Both", "both")


def tech_tags(platform: str, title: str, objective: str, std_ref: str) -> list[str]:
    hay = f"{title} {objective} {std_ref}".lower()
    tech: list[str] = []
    if platform == "web":
      if "graphql" in hay:
          tech.append("graphql")
      if "websocket" in hay or "ws protocol" in hay:
          tech.append("websocket")
      if any(token in hay for token in ("oauth", "oidc", "pkce", "sso", "magic link")):
          tech.append("oauth")
    if platform == "mobile":
      if any(token in hay for token in ("flutter", "dart", "blutter")):
          tech.append("flutter")
      if any(token in hay for token in ("react native", "hermes", "rn bundle")):
          tech.append("reactnative")
      if any(token in hay for token in ("java", "kotlin", "native")):
          tech.append("native")
    if platform == "desktop":
      if any(token in hay for token in ("electron", "node.js", "chromium")):
          tech.append("electron")
      if any(token in hay for token in ("java", "jar", "swing", "fx")):
          tech.append("java")
      if any(token in hay for token in (".net", "wpf", "winforms", "dpapi", "registry", "clr")):
          tech.append("dotnet")
    return sorted(set(tech))


def load_rows(xlsx_path: Path) -> list[dict]:
    workbook = load_workbook(xlsx_path, data_only=True)
    rows: list[dict] = []

    for sheet_name, (platform, section) in SHEET_MAP.items():
        sheet = workbook[sheet_name]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            ref = values[1]
            if not ref:
                continue

            feature_label = values[3].strip() if section == "custom" and values[3] else None
            title = values[4].strip()
            objective = values[5].strip()
            mode = (values[7] if section == "custom" else values[6]) or "Both"
            row_type = ((values[8] if section == "custom" else values[7]) or "Test").strip().lower()
            severity = ((values[9] if section == "custom" else values[8]) or "Medium").strip().lower()
            status = ((values[10] if section == "custom" else values[9]) or "Not Started").strip()

            rows.append(
                {
                    "id": str(ref).strip(),
                    "stdRef": (values[2] or "").strip(),
                    "group": feature_label or (values[3] or "").strip(),
                    "title": title,
                    "objective": objective,
                    "rowType": row_type,
                    "severity": severity,
                    "status": status,
                    "platform": platform,
                    "section": section,
                    "featureKey": slug(platform, feature_label),
                    "featureLabel": feature_label,
                    "access": normalize_access(str(mode).strip()),
                    "source": "excel",
                    "sourceSheet": sheet_name,
                    "sourceRef": str(ref).strip(),
                    "tags": ["feature-specific"] if section == "custom" and feature_label else ["core"],
                    "tech": tech_tags(platform, title, objective, (values[2] or "").strip()),
                }
            )

    return sorted(rows, key=lambda row: row["id"])


def main() -> int:
    parser = argparse.ArgumentParser(description="Import the Excel workbook into normalized checklist catalog JSON.")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Path to the source workbook.")
    parser.add_argument("--out", default=str(DEFAULT_OUT), help="Path to write the imported catalog JSON.")
    args = parser.parse_args()

    rows = load_rows(Path(args.xlsx))
    out_path = Path(args.out)
    out_path.write_text(json.dumps({"rows": rows}, indent=2) + "\n", encoding="utf-8")
    print(f"Imported {len(rows)} Excel rows into {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
